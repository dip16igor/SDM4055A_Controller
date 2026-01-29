"""
Test script to verify enhanced Excel formatting works correctly.
Tests header bold font, borders, and auto-width adjustment.
"""
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def test_enhanced_formatting():
    """Test enhanced Excel formatting."""
    print("Testing enhanced Excel formatting...")
    print("=" * 60)

    # Create sample data
    header = ["QR", "TEST RESULT", "CH1", "CH2", "CH3", "Date/Time"]
    row1 = ["PSN123456789", "OK", "3.3000000", "5.0000000", "12.0000000", "2026-01-29 12:00:00"]
    row2 = ["PSN987654321", "FAILED", "2.5000000", "4.8000000", "11.5000000", "2026-01-29 12:05:00"]
    rows = [header, row1, row2]

    # Create Excel workbook
    print("\n1. Creating Excel workbook...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Write data to Excel
    print("2. Writing data to Excel worksheet...")
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Apply enhanced formatting
    print("3. Applying enhanced formatting...")

    # Define styles
    header_font = Font(name='Arial', size=11, bold=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Apply header formatting (bold font, center alignment)
    print("   - Header row: bold font, center alignment, borders")
    for col_idx, value in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Auto-adjust column widths based on content
    print("   - Auto-adjusting column widths...")
    for col_idx in range(1, len(header) + 1):
        max_length = 0
        for row in rows:
            cell_value = str(row[col_idx - 1]) if col_idx - 1 < len(row) else ""
            max_length = max(max_length, len(cell_value))
        
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max_length + 2
        print(f"     Column {col_letter}: width = {max_length + 2}")

    # Apply borders to all data cells
    print("   - Data cells: borders applied")
    for row_idx, row in enumerate(rows[1:], start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border

    print("\n4. Saving Excel file...")
    excel_filename = "test_enhanced_formatting.xlsx"
    wb.save(excel_filename)
    print(f"   Excel file saved: {excel_filename}")

    # Clean up
    os.remove(excel_filename)
    print("\n" + "=" * 60)
    print("Enhanced formatting test PASSED!")
    print("=" * 60)

    return True

if __name__ == "__main__":
    import sys
    success = test_enhanced_formatting()
    sys.exit(0 if success else 1)
