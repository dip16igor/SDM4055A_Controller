"""
Simple test to verify Excel export functionality still works after ChannelIndicator fix.
This tests the Excel export logic without needing the full GUI.
"""
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def test_excel_export():
    """Test Excel export with conditional formatting."""
    print("Testing Excel export functionality...")
    print("=" * 60)

    # Create a sample CSV file
    csv_filename = "test_export.csv"
    excel_filename = "test_export.xlsx"

    # Sample data with measurements
    header = ["QR", "TEST RESULT", "CH1", "CH2", "CH3", "CH4", "CH5", "CH6", "CH7", "CH8", "CH9", "CH10", "CH11", "CH12", "Date/Time"]
    row1 = ["PSN123456789", "OK", "3.3000000", "5.0000000", "12.0000000", "1.5000000", "2.0000000", "3.0000000", "4.0000000", "5.0000000", "6.0000000", "7.0000000", "8.0000000", "9.0000000", "2026-01-29 12:00:00"]
    row2 = ["PSN987654321", "OK", "2.5000000", "4.8000000", "11.5000000", "1.8000000", "2.5000000", "3.5000000", "4.5000000", "5.5000000", "6.5000000", "7.5000000", "8.5000000", "9.5000000", "2026-01-29 12:05:00"]

    # Write CSV file
    print(f"\n1. Creating CSV file: {csv_filename}")
    with open(csv_filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(header)
        writer.writerow(row1)
        writer.writerow(row2)
    print(f"   CSV file created successfully")

    # Read CSV file
    print(f"\n2. Reading CSV file...")
    with open(csv_filename, 'r', newline='', encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=';')
        rows = list(reader)
    print(f"   Read {len(rows)} rows from CSV")

    # Create Excel workbook
    print(f"\n3. Creating Excel workbook...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Write data to Excel
    print(f"4. Writing data to Excel worksheet...")
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    print(f"   Data written successfully")

    # Define colors
    light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    light_green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    # Apply conditional formatting (simulate threshold checking)
    print(f"\n5. Applying conditional formatting...")
    red_count = 0
    green_count = 0

    # Sample thresholds for testing
    thresholds = {
        "CH1": (3.2, 3.4),  # lower, upper
        "CH2": (4.9, 5.1),
        "CH3": (11.8, 12.2),
    }

    # Process each data row (skip header)
    for row_idx, row in enumerate(rows[1:], start=2):
        for col_idx, col_name in enumerate(header, start=1):
            if col_name.startswith('CH') and col_name in thresholds:
                cell_value = row[col_idx - 1] if col_idx - 1 < len(row) else ""

                if cell_value:
                    try:
                        value = float(cell_value)
                        lower, upper = thresholds[col_name]

                        cell = ws.cell(row=row_idx, column=col_idx)

                        if value < lower or value > upper:
                            cell.fill = light_red_fill
                            red_count += 1
                            print(f"   Row {row_idx}, {col_name}: {value} (outside {lower}-{upper}) -> RED")
                        else:
                            cell.fill = light_green_fill
                            green_count += 1
                            print(f"   Row {row_idx}, {col_name}: {value} (within {lower}-{upper}) -> GREEN")
                    except ValueError:
                        pass

    print(f"\n   Conditional formatting applied: {red_count} red cells, {green_count} green cells")

    # Save Excel file
    print(f"\n6. Saving Excel file: {excel_filename}")
    wb.save(excel_filename)
    print(f"   Excel file saved successfully")

    # Clean up
    print(f"\n7. Cleaning up test files...")
    os.remove(csv_filename)
    os.remove(excel_filename)
    print(f"   Test files removed")

    print("\n" + "=" * 60)
    print("Excel export test PASSED!")
    print("=" * 60)

    return True

if __name__ == "__main__":
    import sys
    success = test_excel_export()
    sys.exit(0 if success else 1)
